from client.client import Client
from .parser import *
from .errors import *
import os
import logging
from pathlib import Path
from openpyxl import load_workbook, Workbook

logger = logging.getLogger(__name__)

POLICY_TARGETS = '6c488338-8eec-4103-ad21-cd461ac2c476'

GROUP_SIZE = 10

def append_list_to_excel(file_name, data_list):
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    ws.append(data_list)
    wb.save(file_name)
    logger.info(f"Appended {data_list[1]} to {file_name}")

def append_row_to_excel(filename, sheet_name, row_data, headers=None):
    # Get user's Desktop path
    desktop_path = Path.home() / "Desktop"
    full_path = desktop_path / filename

    try:
        # Try to load existing workbook
        workbook = load_workbook(full_path)
    except FileNotFoundError:
        # If file doesn't exist, create new one and remove default sheet
        workbook = Workbook()
        workbook.remove(workbook.active)

    # Check if sheet exists
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        # Create new sheet and add headers
        sheet = workbook.create_sheet(title=sheet_name)
        if headers:
            sheet.append(headers)

    # Append data row
    sheet.append(row_data)

    # Save workbook to Desktop
    workbook.save(full_path)

def find_rule_num(unfiltered_rules: list[dict],requested_rule: dict):
    for rule, section in unfiltered_rules:
        if rule['uid'] == requested_rule['uid']:
            return unfiltered_rules.index((rule,section))
    return -1

def get_rulebase(client: Client, rulebase: str, size: int,filter = ''):
    result = []
    offset = 0
    num = size
    while(num > 0):
        if num > GROUP_SIZE:
            rules = client.api_call('show-access-rulebase',{'name':rulebase,'show-hits':True,'limit':GROUP_SIZE,'offset':offset,'filter':filter}).response()['data']['rulebase']
        else:
            rules = client.api_call('show-access-rulebase',{'name':rulebase,'show-hits':True,'limit':num,'offset':offset,'filter':filter}).response()['data']['rulebase']
        for rule in rules:
            if 'rulebase' in rule:
                for sect_rule in rule['rulebase']:
                    result.append((sect_rule,rule['name']))
            else:
                result.append((rule,'NO TITLE'))
        num = num - GROUP_SIZE
        offset = offset + GROUP_SIZE
    return result

def filter_low_activity(client: Client,args: list):
    ##TO DO
    ###ADD IN ARG PARSING
    rule_base = client.api_call('show-access-rulebase',{'name':args[0],'show-hits':True,'limit':1}).response()['data']
    filter = '' if len(args) == 1 else args[1]
    N = rule_base['total']
    filtered_rules = get_rulebase(client,args[0],N,filter)
    unfiltered_rules = get_rulebase(client,args[0],N)
    append_list_to_excel('filtered_list.xlsx', ['Rule Number','Rule Name','Hit Count'])
    for rule, _ in filtered_rules:
        if 'hits' in rule.keys(): 
            if rule['hits']['value'] < 100 and rule['enabled'] == True:
                rule_num = find_rule_num(unfiltered_rules,rule) + 1
                name = '' if 'name' not in rule else rule['name']
                append_list_to_excel('filtered_list.xlsx', [rule_num,name,rule['hits']['value']])

def filter_target(client: Client, args: list):
    rule_base = client.api_call('show-access-rulebase',{'name':args[0],'limit':1}).response()['data']
    #filter = '' if len(args) == 2 else args[2]
    target = args[1]
    N = rule_base['total']
    if target == 'Policy Targets':
        all_rules = get_rulebase(client,args[0],N)
        rules = []
        for rule, section in all_rules:
            if rule['install-on'][0] == POLICY_TARGETS:
                rules.append((rule,section))
    else:
        rules = get_rulebase(client,args[0],N,f'installOn:{target}')
    filename = f'{args[0]}-{client.login_timestamp}.xlsx'
    disabled = 0
    for rule,section in rules:
        rulename = '' if 'name' not in rule else rule['name']
        if rule['enabled'] == False: disabled += 1
        append_row_to_excel(filename,target,[rulename,rule['enabled'],rule['hits']['level'],rule['hits']['value'],section],['Rule Name','Enabled','Activity Level','Hits','Section Title'])
    enabled = len(rules) - disabled
    append_row_to_excel(filename,'Summary',[target,len(rules),enabled,disabled],['Policy Target','Total Rules','Total Enabled Rules','Total Disabled Rules'])
        

            
